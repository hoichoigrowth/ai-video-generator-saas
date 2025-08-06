"""
Data Export Service - Replacing Google Sheets export functionality
CSV, Excel, and PDF export for all project data
"""

import asyncio
import csv
import json
import io
from typing import List, Dict, Any, Optional, Union
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import logging

from database.connection import get_db_session
from database.models import *
from services.storage_service import storage_service
from core.exceptions import AIVideoGeneratorException

logger = logging.getLogger(__name__)

class ExportError(AIVideoGeneratorException):
    """Export service specific errors"""
    pass

class DataExportService:
    """
    Comprehensive data export service
    Replaces Google Sheets with CSV/Excel/PDF exports
    """
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
    
    async def export_shot_division_csv(self, project_id: str) -> Dict[str, Any]:
        """Export shot division data to CSV"""
        try:
            async with get_db_session() as session:
                from sqlalchemy import select
                
                # Get shot division and shots
                result = await session.execute(
                    select(ShotDivision).where(ShotDivision.project_id == project_id)
                )
                shot_division = result.scalar_one_or_none()
                
                if not shot_division:
                    raise ExportError("No shot division found for project")
                
                # Get all shots
                result = await session.execute(
                    select(Shot).where(Shot.shot_division_id == shot_division.id).order_by(Shot.shot_number)
                )
                shots = result.scalars().all()
                
                # Prepare CSV data
                csv_data = []
                headers = [
                    "Shot Number", "Scene Heading", "Description", "Dialogue",
                    "Shot Type", "Camera Angle", "Camera Movement", "Duration (sec)",
                    "Location", "Characters Present", "Props Needed",
                    "Lighting Notes", "Visual Style", "Physics Notes", "Continuity Notes",
                    "Midjourney Prompt", "Selected Image", "Video Status"
                ]
                csv_data.append(headers)
                
                for shot in shots:
                    row = [
                        shot.shot_number,
                        shot.scene_heading or "",
                        shot.description or "",
                        shot.dialogue or "",
                        shot.shot_type.value if shot.shot_type else "",
                        shot.camera_angle or "",
                        shot.camera_movement.value if shot.camera_movement else "",
                        shot.duration_seconds,
                        shot.location or "",
                        ", ".join(shot.characters_present) if shot.characters_present else "",
                        ", ".join(shot.props_needed) if shot.props_needed else "",
                        shot.lighting_notes or "",
                        shot.visual_style or "",
                        shot.physics_notes or "",
                        shot.continuity_notes or "",
                        shot.midjourney_prompt or "",
                        shot.selected_image_path or "",
                        shot.video_status or ""
                    ]
                    csv_data.append(row)
                
                # Generate CSV content
                csv_buffer = io.StringIO()
                writer = csv.writer(csv_buffer)
                writer.writerows(csv_data)
                csv_content = csv_buffer.getvalue().encode('utf-8')
                
                # Store in MinIO
                filename = f"shot_division_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
                storage_result = await storage_service.store_export_file(
                    project_id, csv_content, filename, "csv"
                )
                
                # Update database record
                await session.execute(
                    update(ShotDivision)
                    .where(ShotDivision.id == shot_division.id)
                    .values(csv_export_path=storage_result["file_path"])
                )
                await session.commit()
                
                return {
                    "file_path": storage_result["file_path"],
                    "download_url": storage_result["url"],
                    "file_size": storage_result["size"],
                    "record_count": len(shots),
                    "exported_at": datetime.utcnow().isoformat()
                }
                
        except Exception as e:
            logger.error(f"Failed to export shot division CSV: {e}")
            raise ExportError(f"CSV export failed: {str(e)}")
    
    async def export_shot_division_excel(self, project_id: str) -> Dict[str, Any]:
        """Export shot division data to Excel with formatting"""
        try:
            async with get_db_session() as session:
                from sqlalchemy import select
                
                # Get data
                result = await session.execute(
                    select(ShotDivision).where(ShotDivision.project_id == project_id)
                )
                shot_division = result.scalar_one_or_none()
                
                if not shot_division:
                    raise ExportError("No shot division found for project")
                
                result = await session.execute(
                    select(Shot).where(Shot.shot_division_id == shot_division.id).order_by(Shot.shot_number)
                )
                shots = result.scalars().all()
                
                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Shot Division"
                
                # Styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                center_alignment = Alignment(horizontal="center", vertical="center")
                
                # Headers
                headers = [
                    "Shot #", "Scene Heading", "Description", "Dialogue",
                    "Shot Type", "Camera Angle", "Movement", "Duration",
                    "Location", "Characters", "Props", "Lighting",
                    "Visual Style", "Physics Notes", "Continuity", "MJ Prompt"
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                
                # Data rows
                for row_idx, shot in enumerate(shots, 2):
                    data = [
                        shot.shot_number,
                        shot.scene_heading,
                        shot.description,
                        shot.dialogue,
                        shot.shot_type.value if shot.shot_type else "",
                        shot.camera_angle,
                        shot.camera_movement.value if shot.camera_movement else "",
                        shot.duration_seconds,
                        shot.location,
                        ", ".join(shot.characters_present) if shot.characters_present else "",
                        ", ".join(shot.props_needed) if shot.props_needed else "",
                        shot.lighting_notes,
                        shot.visual_style,
                        shot.physics_notes,
                        shot.continuity_notes,
                        shot.midjourney_prompt[:50] + "..." if shot.midjourney_prompt and len(shot.midjourney_prompt) > 50 else shot.midjourney_prompt
                    ]
                    
                    for col, value in enumerate(data, 1):
                        ws.cell(row=row_idx, column=col, value=value)
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to bytes
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_content = excel_buffer.getvalue()
                
                # Store in MinIO
                filename = f"shot_division_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
                storage_result = await storage_service.store_export_file(
                    project_id, excel_content, filename, "xlsx"
                )
                
                # Update database
                await session.execute(
                    update(ShotDivision)
                    .where(ShotDivision.id == shot_division.id)
                    .values(excel_export_path=storage_result["file_path"])
                )
                await session.commit()
                
                return {
                    "file_path": storage_result["file_path"],
                    "download_url": storage_result["url"],
                    "file_size": storage_result["size"],
                    "record_count": len(shots),
                    "exported_at": datetime.utcnow().isoformat()
                }
                
        except Exception as e:
            logger.error(f"Failed to export shot division Excel: {e}")
            raise ExportError(f"Excel export failed: {str(e)}")
    
    async def export_characters_csv(self, project_id: str) -> Dict[str, Any]:
        """Export character data to CSV"""
        try:
            async with get_db_session() as session:
                from sqlalchemy import select
                
                # Get characters
                result = await session.execute(
                    select(Character).where(Character.project_id == project_id)
                )
                characters = result.scalars().all()
                
                if not characters:
                    raise ExportError("No characters found for project")
                
                # Prepare CSV data
                csv_data = []
                headers = [
                    "Name", "Importance Level", "Description", "Age", "Gender",
                    "Height", "Build", "Hair Color", "Hair Style", "Eye Color",
                    "Skin Tone", "Distinctive Features", "Clothing Style",
                    "Personality Traits", "First Appearance", "Total Scenes",
                    "Midjourney Prompt", "Selected Image", "Approval Status"
                ]
                csv_data.append(headers)
                
                for character in characters:
                    row = [
                        character.name,
                        character.importance_level,
                        character.description or "",
                        character.age or "",
                        character.gender or "",
                        character.height or "",
                        character.build or "",
                        character.hair_color or "",
                        character.hair_style or "",
                        character.eye_color or "",
                        character.skin_tone or "",
                        character.distinctive_features or "",
                        character.clothing_style or "",
                        ", ".join(character.personality_traits) if character.personality_traits else "",
                        character.first_appearance_scene or "",
                        character.total_scenes,
                        character.midjourney_prompt or "",
                        character.selected_image_path or "",
                        character.approval_status.value if character.approval_status else ""
                    ]
                    csv_data.append(row)
                
                # Generate CSV
                csv_buffer = io.StringIO()
                writer = csv.writer(csv_buffer)
                writer.writerows(csv_data)
                csv_content = csv_buffer.getvalue().encode('utf-8')
                
                # Store in MinIO
                filename = f"characters_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
                storage_result = await storage_service.store_export_file(
                    project_id, csv_content, filename, "csv"
                )
                
                return {
                    "file_path": storage_result["file_path"],
                    "download_url": storage_result["url"],
                    "file_size": storage_result["size"],
                    "record_count": len(characters),
                    "exported_at": datetime.utcnow().isoformat()
                }
                
        except Exception as e:
            logger.error(f"Failed to export characters CSV: {e}")
            raise ExportError(f"Characters CSV export failed: {str(e)}")
    
    async def export_production_plan_pdf(self, project_id: str) -> Dict[str, Any]:
        """Export production plan to PDF"""
        try:
            async with get_db_session() as session:
                from sqlalchemy import select
                
                # Get production plan
                result = await session.execute(
                    select(ProductionPlan).where(ProductionPlan.project_id == project_id)
                )
                production_plan = result.scalar_one_or_none()
                
                if not production_plan:
                    raise ExportError("No production plan found for project")
                
                # Get project info
                result = await session.execute(
                    select(Project).where(Project.id == project_id)
                )
                project = result.scalar_one_or_none()
                
                # Create PDF
                pdf_buffer = io.BytesIO()
                doc = SimpleDocTemplate(
                    pdf_buffer,
                    pagesize=A4,
                    rightMargin=72,
                    leftMargin=72,
                    topMargin=72,
                    bottomMargin=72
                )
                
                # Styles
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=self.styles['Heading1'],
                    fontSize=20,
                    spaceAfter=30,
                    alignment=1  # Center
                )
                
                heading_style = ParagraphStyle(
                    'CustomHeading',
                    parent=self.styles['Heading2'],
                    fontSize=14,
                    spaceAfter=12
                )
                
                # Content
                story = []
                
                # Title
                story.append(Paragraph(f"Production Plan: {project.name if project else 'Unknown Project'}", title_style))
                story.append(Spacer(1, 20))
                
                # Project Overview
                story.append(Paragraph("Project Overview", heading_style))
                overview_data = [
                    ["Project ID:", str(project_id)],
                    ["Generated:", datetime.utcnow().strftime("%Y-%m-%d %H:%M")],
                    ["Visual Style:", production_plan.visual_style or "Not specified"],
                    ["Aspect Ratio:", production_plan.aspect_ratio or "9:16"],
                    ["Resolution:", production_plan.resolution or "1080x1920"],
                    ["Frame Rate:", production_plan.frame_rate or "24fps"]
                ]
                
                overview_table = Table(overview_data)
                overview_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ]))
                story.append(overview_table)
                story.append(Spacer(1, 20))
                
                # Timeline
                story.append(Paragraph("Timeline", heading_style))
                timeline_data = [
                    ["Phase", "Duration (days)"],
                    ["Pre-production", str(production_plan.pre_production_days)],
                    ["Production", str(production_plan.production_days)],
                    ["Post-production", str(production_plan.post_production_days)],
                    ["Total", str(production_plan.timeline_days or "TBD")]
                ]
                
                timeline_table = Table(timeline_data)
                timeline_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(timeline_table)
                story.append(Spacer(1, 20))
                
                # Budget
                if production_plan.estimated_budget:
                    story.append(Paragraph("Budget Estimate", heading_style))
                    story.append(Paragraph(f"${production_plan.estimated_budget:,.2f}", self.styles['Normal']))
                    story.append(Spacer(1, 20))
                
                # Quality Standards
                story.append(Paragraph("Quality Standards", heading_style))
                quality_data = [
                    ["Visual Consistency:", production_plan.visual_consistency or "High"],
                    ["Character Continuity:", production_plan.character_continuity or "Strict"],
                    ["Physics Realism:", production_plan.physics_realism or "High"]
                ]
                
                for item in quality_data:
                    story.append(Paragraph(f"<b>{item[0]}</b> {item[1]}", self.styles['Normal']))
                
                story.append(Spacer(1, 20))
                
                # Locations
                if production_plan.locations:
                    story.append(Paragraph("Locations", heading_style))
                    for location, details in production_plan.locations.items():
                        if isinstance(details, dict):
                            story.append(Paragraph(f"<b>{location}</b>: {details.get('description', 'No description')}", self.styles['Normal']))
                        else:
                            story.append(Paragraph(f"<b>{location}</b>", self.styles['Normal']))
                
                # Risk Assessment
                if production_plan.risk_assessment:
                    story.append(Spacer(1, 20))
                    story.append(Paragraph("Risk Assessment", heading_style))
                    for risk in production_plan.risk_assessment:
                        story.append(Paragraph(f"• {risk}", self.styles['Normal']))
                
                # Build PDF
                doc.build(story)
                pdf_content = pdf_buffer.getvalue()
                
                # Store in MinIO
                filename = f"production_plan_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
                storage_result = await storage_service.store_export_file(
                    project_id, pdf_content, filename, "pdf"
                )
                
                # Update database
                await session.execute(
                    update(ProductionPlan)
                    .where(ProductionPlan.id == production_plan.id)
                    .values(pdf_export_path=storage_result["file_path"])
                )
                await session.commit()
                
                return {
                    "file_path": storage_result["file_path"],
                    "download_url": storage_result["url"],
                    "file_size": storage_result["size"],
                    "exported_at": datetime.utcnow().isoformat()
                }
                
        except Exception as e:
            logger.error(f"Failed to export production plan PDF: {e}")
            raise ExportError(f"PDF export failed: {str(e)}")
    
    async def export_project_summary_json(self, project_id: str) -> Dict[str, Any]:
        """Export complete project data as JSON"""
        try:
            async with get_db_session() as session:
                from sqlalchemy import select
                
                # Get all project data
                project_data = {}
                
                # Project info
                result = await session.execute(
                    select(Project).where(Project.id == project_id)
                )
                project = result.scalar_one_or_none()
                
                if project:
                    project_data["project"] = {
                        "id": str(project.id),
                        "name": project.name,
                        "description": project.description,
                        "status": project.status.value,
                        "current_stage": project.current_stage.value,
                        "created_at": project.created_at.isoformat(),
                        "updated_at": project.updated_at.isoformat(),
                        "settings": project.settings,
                        "metadata": project.metadata
                    }
                
                # Screenplays
                result = await session.execute(
                    select(Screenplay).where(Screenplay.project_id == project_id)
                )
                screenplays = result.scalars().all()
                project_data["screenplays"] = [
                    {
                        "id": str(s.id),
                        "version": s.version,
                        "is_current": s.is_current_version,
                        "content_preview": s.content[:200] + "..." if s.content and len(s.content) > 200 else s.content,
                        "approval_status": s.approval_status.value if s.approval_status else None,
                        "quality_score": s.quality_score,
                        "created_at": s.created_at.isoformat()
                    } for s in screenplays
                ]
                
                # Characters
                result = await session.execute(
                    select(Character).where(Character.project_id == project_id)
                )
                characters = result.scalars().all()
                project_data["characters"] = [
                    {
                        "id": str(c.id),
                        "name": c.name,
                        "importance_level": c.importance_level,
                        "description": c.description,
                        "physical_attributes": {
                            "age": c.age,
                            "gender": c.gender,
                            "height": c.height,
                            "build": c.build,
                            "hair_color": c.hair_color,
                            "eye_color": c.eye_color,
                            "skin_tone": c.skin_tone
                        },
                        "personality_traits": c.personality_traits,
                        "approval_status": c.approval_status.value if c.approval_status else None
                    } for c in characters
                ]
                
                # Shot divisions
                result = await session.execute(
                    select(ShotDivision).where(ShotDivision.project_id == project_id)
                )
                shot_divisions = result.scalars().all()
                
                for shot_division in shot_divisions:
                    result = await session.execute(
                        select(Shot).where(Shot.shot_division_id == shot_division.id).order_by(Shot.shot_number)
                    )
                    shots = result.scalars().all()
                    
                    project_data["shot_division"] = {
                        "id": str(shot_division.id),
                        "total_shots": shot_division.total_shots,
                        "estimated_duration": shot_division.estimated_duration,
                        "shots": [
                            {
                                "shot_number": shot.shot_number,
                                "scene_heading": shot.scene_heading,
                                "description": shot.description,
                                "shot_type": shot.shot_type.value if shot.shot_type else None,
                                "duration": shot.duration_seconds,
                                "location": shot.location,
                                "characters": shot.characters_present,
                                "video_status": shot.video_status
                            } for shot in shots
                        ]
                    }
                
                # Convert to JSON
                json_content = json.dumps(project_data, indent=2, default=str).encode('utf-8')
                
                # Store in MinIO
                filename = f"project_summary_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
                storage_result = await storage_service.store_export_file(
                    project_id, json_content, filename, "json"
                )
                
                return {
                    "file_path": storage_result["file_path"],
                    "download_url": storage_result["url"],
                    "file_size": storage_result["size"],
                    "exported_at": datetime.utcnow().isoformat()
                }
                
        except Exception as e:
            logger.error(f"Failed to export project JSON: {e}")
            raise ExportError(f"JSON export failed: {str(e)}")
    
    async def get_export_history(self, project_id: str) -> List[Dict[str, Any]]:
        """Get export history for a project"""
        try:
            async with get_db_session() as session:
                from sqlalchemy import select, desc
                
                result = await session.execute(
                    select(DataExport)
                    .where(DataExport.project_id == project_id)
                    .order_by(desc(DataExport.created_at))
                    .limit(50)
                )
                
                exports = result.scalars().all()
                
                return [
                    {
                        "id": str(export.id),
                        "export_type": export.export_type,
                        "data_type": export.data_type,
                        "file_size": export.file_size,
                        "record_count": export.record_count,
                        "download_count": export.download_count,
                        "created_at": export.created_at.isoformat(),
                        "download_url": storage_service._get_presigned_url(export.file_path) if export.file_path else None
                    } for export in exports
                ]
                
        except Exception as e:
            logger.error(f"Failed to get export history: {e}")
            return []

# Global export service instance
export_service = DataExportService()

# Helper functions for common exports
async def export_project_shots_csv(project_id: str) -> Dict[str, Any]:
    """Helper to export shot division as CSV"""
    return await export_service.export_shot_division_csv(project_id)

async def export_project_shots_excel(project_id: str) -> Dict[str, Any]:
    """Helper to export shot division as Excel"""
    return await export_service.export_shot_division_excel(project_id)

async def export_project_characters(project_id: str) -> Dict[str, Any]:
    """Helper to export characters as CSV"""
    return await export_service.export_characters_csv(project_id)

async def export_production_plan(project_id: str) -> Dict[str, Any]:
    """Helper to export production plan as PDF"""
    return await export_service.export_production_plan_pdf(project_id)